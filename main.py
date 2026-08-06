import pandas as pd
import requests
import json
import re
from datetime import date # Added for sheet naming
import openpyxl # Added for direct Excel file manipulation
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill # Added PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import random # Added for selecting remarks

import os
import shutil

INPUT_FOLDER = "input"
OUTPUT_FOLDER = "output"

INPUT_FILE = os.path.join(INPUT_FOLDER, "students.xlsx")

LEETCODE_API_URL = "https://leetcode.com/graphql/"

REQUEST_TIMEOUT = 10

DATE_FORMAT = "%Y-%m-%d"

uploaded_filename = INPUT_FILE

def extract_leetcode_username(link):
    """Extracts the LeetCode username from a profile link."""
    if pd.isna(link) or not isinstance(link, str):
        return None
    match = re.search(r'leetcode\.com/(u/)?([^/]+)', link)
    if match:
        return match.group(2)
    return None
try:
    df = pd.read_excel(uploaded_filename)
    print(f"Successfully loaded '{uploaded_filename}'.")
    print("Original DataFrame head:")
    print(df.head())
except Exception as e:
    print(f"Error loading Excel file: {e}")
    df = pd.DataFrame() # Create an empty DataFrame to avoid errors later
def get_leetcode_solved_problems(username):
    """Fetches the total number of problems solved by a LeetCode user, by difficulty."""
    if not username:
        return {'Easy': None, 'Medium': None, 'Hard': None, 'Total': None}

    headers = {
        'Content-Type': 'application/json'
    }
    query = """
    query userProfileUserQuestionProgress($username: String!) {
      matchedUser(username: $username) {
        submitStats {
          acSubmissionNum {
            difficulty
            count
          }
        }
      }
    }
    """
    variables = {
        'username': username
    }
    payload = {
        'query': query,
        'variables': variables
    }

    try:
        response = requests.post(LEETCODE_API_URL,headers=headers,data=json.dumps(payload),timeout=REQUEST_TIMEOUT)
        response.raise_for_status() # Raise an HTTPError for bad responses (4xx or 5xx)
        data = response.json()

        difficulty_counts = {'Easy': 0, 'Medium': 0, 'Hard': 0, 'Total': 0}

        if data and 'data' in data and 'matchedUser' in data['data'] and data['data']['matchedUser'] and data['data']['matchedUser']['submitStats']:
            submit_stats = data['data']['matchedUser']['submitStats']
            if submit_stats and 'acSubmissionNum' in submit_stats:
                for entry in submit_stats['acSubmissionNum']:
                    diff = entry['difficulty']
                    count = entry['count']
                    if diff == 'Easy':
                        difficulty_counts['Easy'] = count
                    elif diff == 'Medium':
                        difficulty_counts['Medium'] = count
                    elif diff == 'Hard':
                        difficulty_counts['Hard'] = count
                    elif diff == 'All': # This will be our 'Total' if available
                        difficulty_counts['Total'] = count
                return difficulty_counts
        # If user not found or no stats, return default with 0s
        return {'Easy': 0, 'Medium': 0, 'Hard': 0, 'Total': 0}
    except requests.exceptions.RequestException as e:
        print(f"Network or API error for user {username}: {e}")
        return {'Easy': None, 'Medium': None, 'Hard': None, 'Total': None}
    except Exception as e:
        print(f"An unexpected error occurred for user {username}: {e}")
        return {'Easy': None, 'Medium': None, 'Hard': None, 'Total': None}
if not df.empty:
    df['Leetcode Username'] = df['Leetcode Link'].apply(extract_leetcode_username)

    leetcode_stats = []
    # Use a dictionary to store stats for unique usernames to avoid redundant API calls
    processed_username_stats = {}

    for index, row in df.iterrows():
        username = row['Leetcode Username']
        if username and username not in processed_username_stats:
            solved_counts = get_leetcode_solved_problems(username)
            processed_username_stats[username] = solved_counts

    # Convert the processed_username_stats dictionary to a list of dictionaries for DataFrame creation
    for username, counts in processed_username_stats.items():
        leetcode_stats.append({
            'Leetcode Username': username,
            'Easy': counts.get('Easy', 0),
            'Medium': counts.get('Medium', 0),
            'Hard': counts.get('Hard', 0),
            'Total': counts.get('Total', 0)
        })

    leetcode_results_df = pd.DataFrame(leetcode_stats)
    print("\nLeetCode Solved Problems DataFrame head:")
    print(leetcode_results_df.head())
else:
    print("DataFrame is empty, cannot process LeetCode data.")
    leetcode_results_df = pd.DataFrame()


if not df.empty and not leetcode_results_df.empty:
    try:
        # Load the existing workbook using openpyxl first to access all sheets
        workbook = openpyxl.load_workbook(uploaded_filename)

        # --- Find the most recent previous date sheet for comparison ---
        previous_date_sheets = []
        today = date.today()
        for sheet_name in workbook.sheetnames:
            try:
                sheet_date = datetime.strptime(sheet_name, '%Y-%m-%d').date()
                if sheet_date < today:
                    previous_date_sheets.append((sheet_date, sheet_name))
            except ValueError:
                # Ignore sheets that are not named with the YYYY-MM-DD format
                pass

        most_recent_prev_sheet_name = None
        if previous_date_sheets:
            most_recent_prev_sheet_name = max(previous_date_sheets)[1]
            print(f"Found most recent previous date sheet: '{most_recent_prev_sheet_name}'")

        prev_data_df = pd.DataFrame()
        if most_recent_prev_sheet_name:
            prev_sheet = workbook[most_recent_prev_sheet_name]
            prev_data = []
            for row in prev_sheet.iter_rows(min_row=1, values_only=True):
                prev_data.append(row)
            if prev_data:
                prev_data_df = pd.DataFrame(prev_data[1:], columns=prev_data[0])
                prev_data_df['Register Number'] = prev_data_df['Register Number'].astype(str)
                print("Previous data head:")
                print(prev_data_df[['Register Number', 'Total Problems Solved']].head())

        # Merge original data with LeetCode solved problems, including 'YEAR / DEPT / SEC'
        merged_df = pd.merge(df[['REGISTER NUMBER', 'NAME', 'Leetcode Username', 'YEAR / DEPT / SEC']],
                             leetcode_results_df,
                             on='Leetcode Username',
                             how='left')

        # Prepare data for the new sheet according to the requested columns
        output_data_for_sheet = pd.DataFrame()
        output_data_for_sheet['S.No'] = range(1, len(merged_df) + 1)
        output_data_for_sheet['Register Number'] = merged_df['REGISTER NUMBER'].astype(str)
        output_data_for_sheet['Name'] = merged_df['NAME']
        output_data_for_sheet['YEAR / DEPT / SEC'] = merged_df['YEAR / DEPT / SEC'] # Added new column
        output_data_for_sheet['Easy'] = merged_df['Easy'].fillna(0).astype(int)
        output_data_for_sheet['Medium'] = merged_df['Medium'].fillna(0).astype(int)
        output_data_for_sheet['Hard'] = merged_df['Hard'].fillna(0).astype(int)
        output_data_for_sheet['Total Problems Solved'] = merged_df['Total'].fillna(0).astype(int)

        # --- Calculate Performance Compared to Previous Date ---
        performance_col = []
        if not prev_data_df.empty:
            for index, row in output_data_for_sheet.iterrows():
                reg_num = str(row['Register Number'])
                current_total = row['Total Problems Solved']

                prev_record = prev_data_df[prev_data_df['Register Number'] == reg_num]
                if not prev_record.empty:
                    previous_total = int(prev_record['Total Problems Solved'].iloc[0])
                    difference = current_total - previous_total
                    if difference > 0:
                        performance_col.append(f'Improved (+{difference})')
                    elif difference < 0:
                        performance_col.append(f'Declined ({difference})')
                    else:
                        performance_col.append('No Change (0)')
                else:
                    performance_col.append('New Student')
        else:
            performance_col = ['New Student'] * len(output_data_for_sheet)

        output_data_for_sheet['Performance Compared to Previous Date'] = performance_col

        # --- Add Remarks Column ---
        remarks_zero = [
            "Solve problems consistently.",
            "Practicing daily to build consistency.",
            "Solve at least one problem today.",
            "Regular practice will improve your coding skills.",
            "Make problem-solving a daily habit."
        ]

        remarks_low = [
            "Good effort! Maintain your consistency.",
            "Nice start! Keep solving regularly.",
            "Good progress. Stay consistent.",
            "Keep up the good work and solve problems consistently.",
            "Well done! Continue practicing every day."
        ]

        remarks_high = [
            "Great job! Remember, consistency in practice is key to long-term success.",
            "Excellent progress! Keep building on this momentum, consistent effort always pays off.",
            "Outstanding work! Maintain your consistent practice to truly master problem-solving.",
            "Impressive number of problems solved! Continue with your consistent efforts.",
            "Fantastic performance! Remember, sustained practice is more valuable than sporadic bursts of effort."
        ]

        def get_remark(row):
            total_solved = row['Total Problems Solved']
            performance_status = row['Performance Compared to Previous Date']

            if performance_status == 'No Change (0)':
                return "" # No remark for 'No Change (0)'
            elif total_solved == 0:
                return random.choice(remarks_zero)
            elif 1 <= total_solved <= 5:
                return random.choice(remarks_low)
            else: # total_solved >= 6
                return random.choice(remarks_high)

        output_data_for_sheet['Remarks'] = output_data_for_sheet.apply(get_remark, axis=1)

        # Get today's date for the new sheet name
        today_date_str = date.today().strftime(DATE_FORMAT) # YYYY-MM-DD format

        # Create a new sheet with today's date as the name
        new_sheet_name_base = today_date_str
        new_sheet_name = new_sheet_name_base
        counter = 1
        while new_sheet_name in workbook.sheetnames:
            new_sheet_name = f"{new_sheet_name_base}_{counter}"
            counter += 1

        new_sheet = workbook.create_sheet(new_sheet_name)

        # Write headers and data to the new sheet
        headers = output_data_for_sheet.columns.tolist()
        for r_idx, row_data in enumerate(dataframe_to_rows(output_data_for_sheet, index=False, header=True)):
            new_sheet.append(row_data)

        # --- Apply Formatting ---
        thin_border = Border(left=Side(style='thin', color='FF000000'),
                             right=Side(style='thin', color='FF000000'),
                             top=Side(style='thin', color='FF000000'),
                             bottom=Side(style='thin', color='FF000000'))

        # Get column indices for easy access
        col_indices = {col_name: idx + 1 for idx, col_name in enumerate(headers)}

        # 1. Header Row Formatting (Bold, Center-aligned, Borders, Wrap Text)
        for col_idx, cell in enumerate(new_sheet[1], 1):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            cell.border = thin_border

        # 2. Apply general formatting (Wrap Text, Borders) and specific alignments
        # Iterate through all populated cells in the new sheet
        for row_idx in range(1, new_sheet.max_row + 1):
            for col_idx in range(1, new_sheet.max_column + 1):
                cell = new_sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None: # Apply only to populated cells
                    cell.alignment = Alignment(wrapText=True, horizontal=cell.alignment.horizontal, vertical='center') # Set vertical to center for all data cells
                    cell.border = thin_border

        # 3. Format 'Register Number' column as Number with 0 decimal places
        for cell in new_sheet[get_column_letter(col_indices['Register Number'])][1:]: # Skip header
            if cell.value is not None:
                cell.number_format = '0'

        # 4. Center-align specific columns (data rows, and headers are already centered)
        center_align_cols = ['Register Number', 'Easy', 'Medium', 'Hard', 'Total Problems Solved', 'Performance Compared to Previous Date']
        # Add 'Remarks' to the center-aligned columns if desired, otherwise it will default to general alignment
        # For remarks, typically left-aligned is better, so not adding it here.
        for col_name in center_align_cols:
            if col_name in col_indices:
                col_letter = get_column_letter(col_indices[col_name])
                for row_idx in range(1, new_sheet.max_row + 1):
                    cell = new_sheet.cell(row=row_idx, column=col_indices[col_name])
                    if cell.value is not None:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

        # NEW: Conditional Formatting for 'Performance Compared to Previous Date'
        dark_green_fill = PatternFill(start_color="00008000", end_color="00008000", fill_type="solid") # Dark Green
        light_green_fill = PatternFill(start_color="0090EE90", end_color="0090EE90", fill_type="solid") # Light Green
        dull_gray_fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid") # Dull Gray
        red_fill = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type="solid") # Red Fill

        if 'Performance Compared to Previous Date' in col_indices:
            performance_col_idx = col_indices['Performance Compared to Previous Date']
            for row_idx in range(2, new_sheet.max_row + 1): # Start from row 2 to skip header
                cell = new_sheet.cell(row=row_idx, column=performance_col_idx)
                if cell.value is not None and isinstance(cell.value, str):
                    if cell.value.startswith('Improved (+'):
                        match = re.search(r'Improved \(\+(\d+)\)', cell.value)
                        if match:
                            improved_count = int(match.group(1))
                            if improved_count > 20:
                                cell.fill = dark_green_fill
                                cell.font = Font(color="FFFFFFFF") # White text for dark background
                            elif improved_count > 10: # 10 < diff <= 20
                                cell.fill = light_green_fill
                            elif improved_count > 0: # 0 < diff <= 10
                                cell.fill = dull_gray_fill
                            else: # Fallback for improved_count <= 0, though not expected here
                                cell.fill = dull_gray_fill
                        else:
                            cell.fill = dull_gray_fill # Fallback if pattern doesn't match
                    elif cell.value == 'New Student':
                        cell.fill = light_green_fill
                    elif cell.value == 'No Change (0)': # Specific condition for No Change (0)
                        cell.fill = red_fill
                    elif cell.value.startswith('Declined ('): # Declined still dull gray
                        cell.fill = dull_gray_fill

        # 7. Auto-adjust column widths for readability
        for col_idx, column_title in enumerate(headers, 1):
            max_length = 0
            for row_idx in range(1, new_sheet.max_row + 1): # Iterate all rows
                cell_value = new_sheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    # Heuristics for width: consider number of chars, and if it's a long text, it will wrap.
                    # For now, just measure the longest string representation.
                    length = len(str(cell_value))
                    max_length = max(max_length, length)

            adjusted_width = (max_length + 2) # Add padding
            if column_title == 'Name': # Special handling for Name if needed, but general auto-adjust should be fine
                 adjusted_width = max(adjusted_width, 20) # Ensure a minimum width for names
            elif column_title == 'YEAR / DEPT / SEC': # Ensure minimum width for new column
                 adjusted_width = max(adjusted_width, 15) # Example minimum width
            elif column_title == 'Remarks': # Ensure minimum width for remarks
                 adjusted_width = max(adjusted_width, 40) # Example minimum width for remarks
            new_sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # ----------------------------------------------------
        # Save the workbook once
        # ----------------------------------------------------

        os.makedirs(OUTPUT_FOLDER, exist_ok=True)

        temp_file = os.path.join(INPUT_FOLDER, "students_new.xlsx")

        workbook.save(temp_file)

        # ----------------------------------------------------
        # Save dated copy in output folder
        # ----------------------------------------------------

        output_filename = f"{today_date_str}.xlsx"
        output_file = os.path.join(OUTPUT_FOLDER, output_filename)

        shutil.copy2(temp_file, output_file)

        print(f"Saved report: {output_file}")

        # ----------------------------------------------------
        # Replace input/students.xlsx
        # ----------------------------------------------------

        os.replace(temp_file, INPUT_FILE)

        print("Updated input/students.xlsx")
    except Exception as e:
        print(f"Error processing or modifying the Excel file: {e}")
        import traceback
        traceback.print_exc()
        print("No Excel file was updated in place due to the error.")
else:
    print("DataFrame is empty or no LeetCode statistics were retrieved. No Excel file was updated.")