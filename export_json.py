import openpyxl
import pandas as pd
import json
import re
import os
import glob
from datetime import datetime, date, timedelta

INPUT_FILE = os.path.join("input", "students.xlsx")
OUTPUT_JSON_PATH = os.path.join("output", "leetcode-data.json")

# Primary path to public folder in React dashboard (inside repository)
DASHBOARD_JSON_PATH = os.path.join("leetcode-dashboard", "public", "data", "leetcode-data.json")
# Sibling fallback path for local development
SIBLING_DASHBOARD_JSON_PATH = os.path.normpath(
    os.path.join("..", "leetcode-dashboard", "public", "data", "leetcode-data.json")
)

DATA_DIR = "data"
SUBMISSIONS_DIR = os.path.join(DATA_DIR, "submissions")
CACHE_FILE = os.path.join(DATA_DIR, "problem-cache.json")

def extract_leetcode_username(link):
    if pd.isna(link) or not isinstance(link, str):
        return ""
    match = re.search(r'leetcode\.com/(u/)?([^/]+)', link)
    if match:
        return match.group(2)
    return ""

def parse_sheet_date(sheet_name):
    match = re.match(r'^(\d{4}-\d{2}-\d{2})', sheet_name)
    if match:
        try:
            return datetime.strptime(match.group(1), '%Y-%m-%d').date()
        except ValueError:
            return None
    return None

def format_date_display(dt):
    if isinstance(dt, (datetime, date)):
        return dt.strftime('%d %b')
    return str(dt)

def compute_submission_and_topic_analytics():
    cache = {}
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                cache = json.load(f)
        except Exception as e:
            print(f"Warning: Failed to load problem cache in export_json ({e})")

    sub_files = glob.glob(os.path.join(SUBMISSIONS_DIR, "*.json"))
    
    total_accepted_submissions = 0
    students_with_data = set()
    unique_slugs_solved = set()
    latest_collection_date = ""

    topic_overall = {}
    topic_by_class = {}
    topic_by_diff = {}

    diff_counts = {"Easy": 0, "Medium": 0, "Hard": 0}
    diff_by_topic = {}

    seen_student_problems = set()

    for sfile in sorted(sub_files):
        fname = os.path.basename(sfile).replace(".json", "")
        if fname > latest_collection_date:
            latest_collection_date = fname
        try:
            with open(sfile, "r", encoding="utf-8") as f:
                data = json.load(f)
                students = data.get("students", {})
                for reg, sdata in students.items():
                    dept = sdata.get("department", "CSD")
                    subs = sdata.get("recentAcceptedSubmissions", [])
                    if subs:
                        students_with_data.add(reg)
                    for sub in subs:
                        total_accepted_submissions += 1
                        tslug = sub.get("titleSlug")
                        if not tslug:
                            continue
                        unique_slugs_solved.add(tslug)

                        # Deduplicate per student per problem slug for topic/difficulty counts
                        dedup_key = (reg, tslug)
                        if dedup_key in seen_student_problems:
                            continue
                        seen_student_problems.add(dedup_key)

                        q_info = cache.get(tslug, {})
                        diff = q_info.get("difficulty", "Unknown")
                        tags = q_info.get("topicTags", [])

                        if diff in diff_counts:
                            diff_counts[diff] += 1

                        for tag in tags:
                            topic_overall[tag] = topic_overall.get(tag, 0) + 1
                            
                            if dept not in topic_by_class:
                                topic_by_class[dept] = {}
                            topic_by_class[dept][tag] = topic_by_class[dept].get(tag, 0) + 1

                            if diff != "Unknown":
                                if diff not in topic_by_diff:
                                    topic_by_diff[diff] = {}
                                topic_by_diff[diff][tag] = topic_by_diff[diff].get(tag, 0) + 1

                                if tag not in diff_by_topic:
                                    diff_by_topic[tag] = {"Easy": 0, "Medium": 0, "Hard": 0}
                                if diff in diff_by_topic[tag]:
                                    diff_by_topic[tag][diff] += 1
        except Exception as e:
            print(f"Warning: Failed reading submission file {sfile}: {e}")

    overall_list = [{"topic": k, "count": v} for k, v in sorted(topic_overall.items(), key=lambda x: x[1], reverse=True)]
    
    by_class_dict = {}
    for dept, tdict in topic_by_class.items():
        by_class_dict[dept] = [{"topic": k, "count": v} for k, v in sorted(tdict.items(), key=lambda x: x[1], reverse=True)]

    by_diff_dict = {}
    for diff, tdict in topic_by_diff.items():
        by_diff_dict[diff] = [{"topic": k, "count": v} for k, v in sorted(tdict.items(), key=lambda x: x[1], reverse=True)]

    topics_analysis = {
        "overall": overall_list,
        "byClass": by_class_dict,
        "byDifficulty": by_diff_dict
    }

    difficulty_analysis = {
        "Easy": diff_counts.get("Easy", 0),
        "Medium": diff_counts.get("Medium", 0),
        "Hard": diff_counts.get("Hard", 0),
        "byTopic": diff_by_topic
    }

    submission_analytics = {
        "totalAcceptedSubmissions": total_accepted_submissions,
        "studentsWithSubmissionData": len(students_with_data),
        "uniqueProblemsSolved": len(unique_slugs_solved),
        "latestCollectionDate": latest_collection_date
    }

    return topics_analysis, difficulty_analysis, submission_analytics

def run_export():
    if not os.path.exists(INPUT_FILE):
        print(f"File not found: {INPUT_FILE}")
        return

    print(f"Loading workbook: {INPUT_FILE}")
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)

    # 1. Parse Account Details
    user_map = {}
    classes_set = set()

    if "Account Details" in wb.sheetnames:
        acc_sheet = wb["Account Details"]
        acc_rows = list(acc_sheet.iter_rows(values_only=True))
        if acc_rows:
            headers = [str(h).strip().upper() if h else '' for h in acc_rows[0]]
            reg_idx = -1
            link_idx = -1
            name_idx = -1
            dept_idx = -1

            for idx, h in enumerate(headers):
                if "REGISTER" in h:
                    reg_idx = idx
                elif "LINK" in h:
                    link_idx = idx
                elif "NAME" in h:
                    name_idx = idx
                elif "YEAR" in h or "DEPT" in h or "SEC" in h:
                    dept_idx = idx

            for row in acc_rows[1:]:
                if not row or reg_idx >= len(row) or not row[reg_idx]:
                    continue
                reg_num = str(row[reg_idx]).strip()
                link = str(row[link_idx]).strip() if link_idx < len(row) and row[link_idx] else ""
                username = extract_leetcode_username(link)
                name = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] else ""
                dept = str(row[dept_idx]).strip() if dept_idx < len(row) and row[dept_idx] else "CSD"

                if dept:
                    classes_set.add(dept)

                user_map[reg_num] = {
                    "username": username,
                    "name": name,
                    "dept": dept,
                    "link": link
                }

    # 2. Identify dated sheets
    dated_sheets = []
    for s_name in wb.sheetnames:
        d = parse_sheet_date(s_name)
        if d:
            dated_sheets.append((d, s_name))

    dated_sheets.sort(key=lambda x: (x[0], x[1]))

    if not dated_sheets:
        print("No dated sheets found in workbook.")
        return

    print(f"Found {len(dated_sheets)} dated sheets.")

    # Deduplicate dated sheets by selecting the latest sheet for each unique date
    unique_dated_sheets = {}
    for d, s_name in dated_sheets:
        unique_dated_sheets[d] = s_name

    sorted_dates = sorted(unique_dated_sheets.keys())

    # 3. Process time-series & student progress & daily snapshots
    daily_trend = []
    class_daily_trends = {}  # class -> list of { date, day, solved }
    student_history = {}    # reg_num -> list of { date, total, easy, medium, hard, imp, dept }
    daily_snapshots = {}    # date_str -> { date, dateFormatted, summary, classSummaries, students }

    for d in sorted_dates:
        date_str = d.strftime('%Y-%m-%d')
        date_fmt = format_date_display(d)
        s_name = unique_dated_sheets[d]
        sheet = wb[s_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h).strip() if h else '' for h in rows[0]]

        reg_col = -1
        name_col = -1
        dept_col = -1
        easy_col = -1
        med_col = -1
        hard_col = -1
        total_col = -1
        perf_col = -1

        for idx, h in enumerate(headers):
            h_upper = h.upper()
            if "REGISTER NUMBER" in h_upper or "REGISTER" in h_upper:
                reg_col = idx
            elif "NAME" in h_upper and name_col == -1:
                name_col = idx
            elif "YEAR" in h_upper or "DEPT" in h_upper or "SEC" in h_upper:
                dept_col = idx
            elif "EASY" in h_upper:
                easy_col = idx
            elif "MEDIUM" in h_upper:
                med_col = idx
            elif "HARD" in h_upper:
                hard_col = idx
            elif "TOTAL" in h_upper:
                total_col = idx
            elif "PERFORMANCE" in h_upper:
                perf_col = idx

        day_total_solved = 0
        day_easy_total = 0
        day_medium_total = 0
        day_hard_total = 0
        day_improved_cnt = 0
        day_nochange_cnt = 0
        day_declined_cnt = 0
        day_active_cnt = 0

        class_day_totals = {}  # class_name -> { total, easy, medium, hard, active, improved, nochange, declined }
        day_student_records = []

        for row in rows[1:]:
            if not row or reg_col >= len(row) or row[reg_col] is None:
                continue

            reg_num = str(row[reg_col]).strip()
            row_name = str(row[name_col]).strip() if name_col != -1 and name_col < len(row) and row[name_col] else ""
            row_dept = str(row[dept_col]).strip() if dept_col != -1 and dept_col < len(row) and row[dept_col] else ""

            acc_info = user_map.get(reg_num, {})
            name = row_name or acc_info.get("name", "") or f"Student {reg_num}"
            dept = row_dept or acc_info.get("dept", "CSD")

            if dept:
                classes_set.add(dept)

            easy = int(row[easy_col]) if easy_col < len(row) and isinstance(row[easy_col], (int, float)) else 0
            medium = int(row[med_col]) if med_col < len(row) and isinstance(row[med_col], (int, float)) else 0
            hard = int(row[hard_col]) if hard_col < len(row) and isinstance(row[hard_col], (int, float)) else 0
            total = int(row[total_col]) if total_col < len(row) and isinstance(row[total_col], (int, float)) else (easy + medium + hard)

            # Parse improvement
            imp = 0
            if perf_col < len(row) and row[perf_col]:
                perf_str = str(row[perf_col])
                match = re.search(r'\((\+?\-?\d+)\)', perf_str)
                if match:
                    imp = int(match.group(1))

            day_total_solved += total
            day_easy_total += easy
            day_medium_total += medium
            day_hard_total += hard

            if total > 0:
                day_active_cnt += 1

            if imp > 0:
                day_improved_cnt += 1
            elif imp < 0:
                day_declined_cnt += 1
            else:
                day_nochange_cnt += 1

            # Update class day totals
            if dept not in class_day_totals:
                class_day_totals[dept] = {
                    "total": 0, "easy": 0, "medium": 0, "hard": 0,
                    "active": 0, "improved": 0, "nochange": 0, "declined": 0
                }
            cdt = class_day_totals[dept]
            cdt["total"] += total
            cdt["easy"] += easy
            cdt["medium"] += medium
            cdt["hard"] += hard
            if total > 0: cdt["active"] += 1
            if imp > 0: cdt["improved"] += 1
            elif imp < 0: cdt["declined"] += 1
            else: cdt["nochange"] += 1

            if reg_num not in student_history:
                student_history[reg_num] = []

            student_history[reg_num].append({
                "date": d,
                "total": total,
                "easy": easy,
                "medium": medium,
                "hard": hard,
                "imp": imp,
                "dept": dept
            })

            day_student_records.append({
                "rollNumber": reg_num,
                "name": name,
                "department": dept,
                "easy": easy,
                "medium": medium,
                "hard": hard,
                "totalSolved": total,
                "improvement": imp
            })

        # Sort day students by totalSolved descending for rank assignment
        day_student_records.sort(key=lambda x: x["totalSolved"], reverse=True)
        for r_idx, s_rec in enumerate(day_student_records):
            s_rec["rank"] = r_idx + 1

        daily_trend.append({
            "date": date_str,
            "day": date_fmt,
            "solved": day_total_solved
        })

        # Update class daily trends
        for dept_name, cdt in class_day_totals.items():
            if dept_name not in class_daily_trends:
                class_daily_trends[dept_name] = []
            class_daily_trends[dept_name].append({
                "date": date_str,
                "day": date_fmt,
                "solved": cdt["total"]
            })

        daily_snapshots[date_str] = {
            "date": date_str,
            "dateFormatted": date_fmt,
            "summary": {
                "totalSolved": day_total_solved,
                "easyTotal": day_easy_total,
                "mediumTotal": day_medium_total,
                "hardTotal": day_hard_total,
                "activeStudents": day_active_cnt,
                "improvedCount": day_improved_cnt,
                "noChangeCount": day_nochange_cnt,
                "declinedCount": day_declined_cnt
            },
            "classSummaries": class_day_totals,
            "students": day_student_records
        }

    # 4. Extract Latest Sheet Data & Class Summaries
    latest_date = sorted_dates[-1]
    prev_date = sorted_dates[-2] if len(sorted_dates) >= 2 else sorted_dates[0]

    students_list = []
    total_easy = 0
    total_medium = 0
    total_hard = 0
    total_solved = 0
    solved_today = 0

    class_student_counts = {}
    class_totals = {}

    # Phase 3 Enhancements: Inactivity, Distribution, Rankings, Weekly Improvement
    active_students_cnt = 0
    inactive_recent_cnt = 0
    inactive_long_cnt = 0
    never_active_cnt = 0
    inactive_students_list = []

    sol_dist = {"solved0": 0, "solved1": 0, "solved2": 0, "solved3plus": 0}
    yesterday_top_list = []
    weekly_top_list = []

    weekly_improved_students_cnt = 0
    weekly_active_students_cnt = 0
    weekly_no_imp_students_cnt = 0
    total_weekly_problems = 0

    for reg_num, history in student_history.items():
        latest_record = history[-1]
        t_solved = latest_record["total"]
        e_count = latest_record["easy"]
        m_count = latest_record["medium"]
        h_count = latest_record["hard"]
        c_imp = latest_record["imp"]
        dept = latest_record["dept"]

        acc_info = user_map.get(reg_num, {})
        username = acc_info.get("username", "") or f"user_{reg_num.lower()}"
        name = acc_info.get("name", "") or f"Student {reg_num}"

        total_easy += e_count
        total_medium += m_count
        total_hard += h_count
        total_solved += t_solved

        if c_imp > 0:
            solved_today += c_imp

        # Distribution (Today / Latest Snapshot)
        if c_imp == 0:
            sol_dist["solved0"] += 1
        elif c_imp == 1:
            sol_dist["solved1"] += 1
        elif c_imp == 2:
            sol_dist["solved2"] += 1
        else:
            sol_dist["solved3plus"] += 1

        # Yesterday's Top Students calculation
        if len(history) >= 2:
            prev_record = history[-2]
            prev_imp = prev_record["imp"]
            if prev_imp > 0:
                yesterday_top_list.append({
                    "rollNumber": reg_num,
                    "name": name,
                    "username": username,
                    "department": dept,
                    "problemsSolved": prev_imp
                })

        # Weekly improvement for student
        start_idx = max(0, len(history) - 7)
        w_imp = history[-1]["total"] - history[start_idx]["total"]
        if w_imp > 0:
            weekly_top_list.append({
                "rollNumber": reg_num,
                "name": name,
                "username": username,
                "department": dept,
                "weeklyProblemsSolved": w_imp
            })
            weekly_improved_students_cnt += 1
            total_weekly_problems += w_imp
        else:
            weekly_no_imp_students_cnt += 1

        if t_solved > 0:
            weekly_active_students_cnt += 1

        # Streak calculation
        streak = 0
        for entry in reversed(history):
            if entry["total"] > 0:
                streak += 1
            else:
                break

        # Last active calculation
        last_active_str = "Today"
        if c_imp == 0:
            days_ago = 0
            for entry in reversed(history):
                if entry["imp"] > 0:
                    break
                days_ago += 1
            if days_ago == 0:
                last_active_str = "Today"
            elif days_ago == 1:
                last_active_str = "Yesterday"
            else:
                last_active_str = f"{days_ago} days ago"

        status = "Active" if t_solved > 0 and (c_imp >= 0 or streak > 0) else "Inactive"

        # Inactivity Categorization
        if t_solved == 0:
            never_active_cnt += 1
            inactive_students_list.append({
                "rollNumber": reg_num,
                "name": name,
                "username": username,
                "department": dept,
                "totalSolved": 0,
                "lastActiveDate": "Never",
                "daysSinceLastActivity": None,
                "inactivityType": "NEVER_ACTIVE"
            })
        else:
            last_active_d = None
            prev_t = 0
            for h in history:
                if h["total"] > prev_t or h["imp"] > 0:
                    last_active_d = h["date"]
                prev_t = h["total"]

            if last_active_d:
                days_since_act = (latest_date - last_active_d).days
                last_act_date_str = last_active_d.strftime("%Y-%m-%d")
            else:
                days_since_act = (latest_date - history[0]["date"]).days
                last_act_date_str = history[0]["date"].strftime("%Y-%m-%d")

            if days_since_act <= 7:
                active_students_cnt += 1
            elif 7 < days_since_act <= 14:
                inactive_recent_cnt += 1
                inactive_students_list.append({
                    "rollNumber": reg_num,
                    "name": name,
                    "username": username,
                    "department": dept,
                    "totalSolved": t_solved,
                    "lastActiveDate": last_act_date_str,
                    "daysSinceLastActivity": days_since_act,
                    "inactivityType": "INACTIVE_RECENT"
                })
            else:
                inactive_long_cnt += 1
                inactive_students_list.append({
                    "rollNumber": reg_num,
                    "name": name,
                    "username": username,
                    "department": dept,
                    "totalSolved": t_solved,
                    "lastActiveDate": last_act_date_str,
                    "daysSinceLastActivity": days_since_act,
                    "inactivityType": "INACTIVE_LONG_TERM"
                })

        # Update class totals
        if dept not in class_student_counts:
            class_student_counts[dept] = 0
            class_totals[dept] = {
                "totalStudents": 0, "activeStudents": 0, "totalProblemsSolved": 0,
                "solvedToday": 0, "weeklyImprovement": 0,
                "easyTotal": 0, "mediumTotal": 0, "hardTotal": 0
            }
        class_student_counts[dept] += 1
        ct = class_totals[dept]
        ct["totalStudents"] += 1
        if status == "Active": ct["activeStudents"] += 1
        ct["totalProblemsSolved"] += t_solved
        if c_imp > 0: ct["solvedToday"] += c_imp
        ct["easyTotal"] += e_count
        ct["mediumTotal"] += m_count
        ct["hardTotal"] += h_count

        # Class weekly improvement calculation
        if w_imp > 0:
            ct["weeklyImprovement"] += w_imp

        students_list.append({
            "id": reg_num,
            "rollNumber": reg_num,
            "name": name,
            "department": dept,
            "leetcodeUsername": username,
            "easy": e_count,
            "medium": m_count,
            "hard": h_count,
            "totalSolved": t_solved,
            "improvement": c_imp,
            "streak": streak,
            "lastActive": last_active_str,
            "status": status
        })

    # Sort rankings
    yesterday_top_list.sort(key=lambda x: x["problemsSolved"], reverse=True)
    weekly_top_list.sort(key=lambda x: x["weeklyProblemsSolved"], reverse=True)
    inactive_students_list.sort(key=lambda x: (x["totalSolved"], x["daysSinceLastActivity"] or 9999), reverse=False)

    total_valid_students = len(students_list)
    student_distribution = {
        "solved0": sol_dist["solved0"],
        "solved1": sol_dist["solved1"],
        "solved2": sol_dist["solved2"],
        "solved3plus": sol_dist["solved3plus"],
        "percentages": {
            "solved0": round((sol_dist["solved0"] / total_valid_students) * 100, 2) if total_valid_students else 0,
            "solved1": round((sol_dist["solved1"] / total_valid_students) * 100, 2) if total_valid_students else 0,
            "solved2": round((sol_dist["solved2"] / total_valid_students) * 100, 2) if total_valid_students else 0,
            "solved3plus": round((sol_dist["solved3plus"] / total_valid_students) * 100, 2) if total_valid_students else 0
        }
    }

    weekly_improvement_stats = {
        "problemsSolved": total_weekly_problems,
        "studentsImproved": weekly_improved_students_cnt,
        "studentsActive": weekly_active_students_cnt,
        "studentsNoImprovement": weekly_no_imp_students_cnt
    }

    # Sort students by totalSolved descending
    students_list.sort(key=lambda x: x["totalSolved"], reverse=True)

    # Compute avgSolved for each class
    for dept_name, ct in class_totals.items():
        ct["avgSolved"] = round(ct["totalProblemsSolved"] / ct["totalStudents"]) if ct["totalStudents"] > 0 else 0

    # Sort classes list alphabetically
    classes_list = sorted(list(classes_set))

    # Build Class Top Performers & Class Summaries
    class_top_performers = {}
    for c_name in classes_list:
        class_students = [s for s in students_list if s["department"] == c_name]
        class_students.sort(key=lambda x: x["totalSolved"], reverse=True)
        class_top_performers[c_name] = class_students[:5]

    # Weekly improvement for department
    weekly_improvement = total_weekly_problems

    # 5. Build Top Performers (Department)
    top_performers = students_list[:5]

    # 6. Compute Submission & Topic Analytics from cached submission data
    topics_analysis, difficulty_analysis, submission_analytics = compute_submission_and_topic_analytics()

    # 7. Build Payload
    export_payload = {
        "latestDate": latest_date.strftime('%Y-%m-%d'),
        "latestDateFormatted": format_date_display(latest_date),
        "classes": classes_list,
        "summary": {
            "totalStudents": len(students_list),
            "activeStudents": active_students_cnt,
            "inactiveStudents": inactive_recent_cnt + inactive_long_cnt,
            "neverActiveStudents": never_active_cnt,
            "totalProblemsSolved": total_solved,
            "solvedToday": solved_today,
            "weeklyImprovement": weekly_improvement,
            "easyTotal": total_easy,
            "mediumTotal": total_medium,
            "hardTotal": total_hard,
            "avgSolved": round(total_solved / len(students_list)) if students_list else 0
        },
        "dailyTrend": daily_trend[-7:] if len(daily_trend) >= 7 else daily_trend,
        "fullHistoryTrend": daily_trend,
        "classDailyTrends": class_daily_trends,
        "classSummaries": class_totals,
        "classTopPerformers": class_top_performers,
        "students": students_list,
        "topPerformers": top_performers,
        "dailySnapshots": daily_snapshots,

        # Phase 3 Analytics Extensions
        "inactiveStudents": inactive_students_list,
        "studentDistribution": student_distribution,
        "yesterdayTopStudents": yesterday_top_list,
        "weeklyTopStudents": weekly_top_list,
        "weeklyImprovement": weekly_improvement_stats,
        "topicsAnalysis": topics_analysis,
        "difficultyAnalysis": difficulty_analysis,
        "submissionAnalytics": submission_analytics
    }

    # 8. Write to output files
    os.makedirs(os.path.dirname(OUTPUT_JSON_PATH), exist_ok=True)
    with open(OUTPUT_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(export_payload, f, indent=2)
    print(f"Saved: {OUTPUT_JSON_PATH}")

    dashboard_dir = os.path.dirname(DASHBOARD_JSON_PATH)
    os.makedirs(dashboard_dir, exist_ok=True)
    with open(DASHBOARD_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(export_payload, f, indent=2)
    print(f"Saved: {DASHBOARD_JSON_PATH}")

    sibling_dir = os.path.dirname(SIBLING_DASHBOARD_JSON_PATH)
    if os.path.exists(os.path.dirname(sibling_dir)):
        os.makedirs(sibling_dir, exist_ok=True)
        with open(SIBLING_DASHBOARD_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(export_payload, f, indent=2)
        print(f"Saved: {SIBLING_DASHBOARD_JSON_PATH}")

if __name__ == "__main__":
    run_export()
