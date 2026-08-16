#!/usr/bin/env python3
"""
collect_submissions.py
Phase 1 & Phase 2 implementation for CSD LeetCode Performance Tracker.

Collects recent accepted submissions and user calendar info for students listed
in input/students.xlsx via public LeetCode GraphQL API. Caches problem metadata
in data/problem-cache.json and writes daily snapshots to data/submissions/YYYY-MM-DD.json.
"""

import os
import sys
import json
import time
import re
import argparse
from datetime import datetime, timezone, timedelta
import requests
import openpyxl

# Define explicit IST timezone (+05:30)
IST = timezone(timedelta(hours=5, minutes=30))

LEETCODE_URL = "https://leetcode.com/graphql/"
HEADERS = {
    "Content-Type": "application/json",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

DATA_DIR = "data"
SUBMISSIONS_DIR = os.path.join(DATA_DIR, "submissions")
CACHE_FILE = os.path.join(DATA_DIR, "problem-cache.json")
EXCEL_PATH = os.path.join("input", "students.xlsx")

# GraphQL Queries
QUERY_USER_DATA = """
query getUserData($username: String!, $year: Int) {
  recentAcSubmissionList(username: $username, limit: 20) {
    id
    title
    titleSlug
    timestamp
    lang
  }
  matchedUser(username: $username) {
    userCalendar(year: $year) {
      streak
      totalActiveDays
      submissionCalendar
      activeYears
    }
  }
}
"""

QUERY_QUESTION_DETAILS = """
query getQuestion($titleSlug: String!) {
  question(titleSlug: $titleSlug) {
    questionId
    title
    difficulty
    topicTags {
      name
      slug
    }
  }
}
"""

def extract_leetcode_username(link):
    if not link or not isinstance(link, str):
        return ""
    match = re.search(r'leetcode\.com/(u/)?([^/]+)', link)
    if match:
        return match.group(2)
    return ""

def load_students_from_excel(excel_path=EXCEL_PATH):
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return []
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "Account Details" not in wb.sheetnames:
        print("Error: 'Account Details' sheet not found in Excel.")
        return []
    
    sheet = wb["Account Details"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    
    headers = [str(h).strip().upper() if h else '' for h in rows[0]]
    reg_idx, name_idx, dept_idx, link_idx = -1, -1, -1, -1
    
    for idx, h in enumerate(headers):
        if "REGISTER" in h:
            reg_idx = idx
        elif "NAME" in h and name_idx == -1:
            name_idx = idx
        elif ("YEAR" in h or "DEPT" in h or "SEC" in h) and dept_idx == -1:
            dept_idx = idx
        elif "LINK" in h or "LEETCODE" in h:
            link_idx = idx
            
    students = []
    for row in rows[1:]:
        if not row or reg_idx >= len(row) or not row[reg_idx]:
            continue
        reg_num = str(row[reg_idx]).strip()
        name = str(row[name_idx]).strip() if name_idx != -1 and name_idx < len(row) and row[name_idx] else ""
        dept = str(row[dept_idx]).strip() if dept_idx != -1 and dept_idx < len(row) and row[dept_idx] else "CSD"
        link = str(row[link_idx]).strip() if link_idx != -1 and link_idx < len(row) and row[link_idx] else ""
        username = extract_leetcode_username(link)
        
        if username:
            students.append({
                "registerNumber": reg_num,
                "name": name,
                "department": dept,
                "username": username,
                "link": link
            })
            
    return students

def load_problem_cache():
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"Warning: Failed to load problem cache ({e}). Starting with empty cache.")
    return {}

def save_problem_cache(cache):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2, ensure_ascii=False)

def fetch_problem_details(title_slug, cache):
    if title_slug in cache:
        return cache[title_slug]
    
    try:
        resp = requests.post(
            LEETCODE_URL,
            json={"query": QUERY_QUESTION_DETAILS, "variables": {"titleSlug": title_slug}},
            headers=HEADERS,
            timeout=10
        )
        if resp.status_code == 200:
            data = resp.json().get("data", {})
            q = data.get("question")
            if q:
                topic_tags = [t["name"] for t in q.get("topicTags", []) if isinstance(t, dict) and "name" in t]
                cache_entry = {
                    "questionId": q.get("questionId", ""),
                    "title": q.get("title", ""),
                    "difficulty": q.get("difficulty", ""),
                    "topicTags": topic_tags
                }
                cache[title_slug] = cache_entry
                save_problem_cache(cache)
                time.sleep(0.2)
                return cache_entry
    except Exception as e:
        print(f"Warning: Failed to fetch problem details for '{title_slug}': {e}")
    
    return None

def fetch_user_data(username, current_year):
    payload = {
        "query": QUERY_USER_DATA,
        "variables": {
            "username": username,
            "year": current_year
        }
    }
    resp = requests.post(LEETCODE_URL, json=payload, headers=HEADERS, timeout=12)
    resp.raise_for_status()
    result = resp.json()
    
    if "errors" in result and not result.get("data"):
        raise RuntimeError(f"GraphQL errors: {result['errors']}")
        
    return result.get("data", {})

def main():
    parser = argparse.ArgumentParser(description="Collect submission data and update problem cache.")
    parser.add_argument("--limit", type=int, default=None, help="Limit total students processed (for testing)")
    parser.add_argument("--username", type=str, default=None, help="Process single student username (for testing)")
    args = parser.parse_args()

    start_time = time.time()
    now_ist = datetime.now(IST)
    today_str = now_ist.strftime("%Y-%m-%d")
    current_year = now_ist.year

    os.makedirs(SUBMISSIONS_DIR, exist_ok=True)
    daily_file_path = os.path.join(SUBMISSIONS_DIR, f"{today_str}.json")

    # Load problem cache
    problem_cache = load_problem_cache()

    # Load existing daily dataset if running multiple times on same day
    daily_dataset = {
        "date": today_str,
        "collectedAt": now_ist.isoformat(),
        "students": {}
    }
    if os.path.exists(daily_file_path):
        try:
            with open(daily_file_path, "r", encoding="utf-8") as f:
                loaded = json.load(f)
                if isinstance(loaded, dict) and "students" in loaded:
                    daily_dataset = loaded
                    daily_dataset["collectedAt"] = now_ist.isoformat()
        except Exception as e:
            print(f"Warning: Could not read existing daily file {daily_file_path}: {e}")

    # Load student list
    students_list = load_students_from_excel()
    if not students_list:
        print("No students loaded. Exiting.")
        sys.exit(1)

    if args.username:
        students_list = [s for s in students_list if s["username"].lower() == args.username.lower()]
        if not students_list:
            # Create a test entry for explicit username test if not in excel
            students_list = [{
                "registerNumber": "TEST000",
                "name": f"Test Student ({args.username})",
                "department": "CSD",
                "username": args.username,
                "link": f"https://leetcode.com/{args.username}"
            }]
    elif args.limit and args.limit > 0:
        students_list = students_list[:args.limit]

    total_students = len(students_list)
    print("Collecting submission data...")

    success_count = 0
    fail_count = 0
    total_submissions_collected = 0

    for idx, student in enumerate(students_list, 1):
        username = student["username"]
        reg_num = student["registerNumber"]
        name = student["name"]
        dept = student["department"]

        sys.stdout.write(f"[{idx}/{total_students}] {username}... ")
        sys.stdout.flush()

        try:
            user_data = fetch_user_data(username, current_year)
            
            recent_subs_raw = user_data.get("recentAcSubmissionList") or []
            matched_user = user_data.get("matchedUser")
            
            if matched_user is None and not recent_subs_raw:
                raise RuntimeError("User not found or no public profile data")

            user_calendar = matched_user.get("userCalendar") if matched_user else {}
            if not user_calendar:
                user_calendar = {}

            streak = user_calendar.get("streak") or 0
            total_active_days = user_calendar.get("totalActiveDays") or 0
            raw_sub_cal = user_calendar.get("submissionCalendar") or "{}"
            
            try:
                calendar_counts = json.loads(raw_sub_cal)
            except Exception:
                calendar_counts = {}

            # Process accepted submissions
            processed_subs = []
            
            # Maintain deduplication set by submission ID
            existing_student_entry = daily_dataset["students"].get(reg_num, {})
            existing_subs = existing_student_entry.get("recentAcceptedSubmissions", [])
            seen_ids = {str(sub["id"]) for sub in existing_subs if "id" in sub}

            # Merge existing submissions first
            for sub in existing_subs:
                processed_subs.append(sub)

            for sub in recent_subs_raw:
                sub_id = str(sub.get("id", ""))
                if sub_id and sub_id in seen_ids:
                    continue
                
                if sub_id:
                    seen_ids.add(sub_id)

                ts = int(sub.get("timestamp", 0))
                if ts > 0:
                    sub_dt = datetime.fromtimestamp(ts, tz=IST)
                    sub_date = sub_dt.strftime("%Y-%m-%d")
                    sub_time = sub_dt.strftime("%H:%M:%S")
                else:
                    sub_date = ""
                    sub_time = ""

                title_slug = sub.get("titleSlug", "")
                title = sub.get("title", "")
                lang = sub.get("lang", "")

                sub_obj = {
                    "id": sub_id,
                    "title": title,
                    "titleSlug": title_slug,
                    "timestamp": ts,
                    "date": sub_date,
                    "time": sub_time,
                    "language": lang
                }
                processed_subs.append(sub_obj)

                # Update problem cache for new title slugs
                if title_slug:
                    fetch_problem_details(title_slug, problem_cache)

            # Sort submissions by timestamp descending
            processed_subs.sort(key=lambda x: x.get("timestamp", 0), reverse=True)

            daily_dataset["students"][reg_num] = {
                "username": username,
                "name": name,
                "department": dept,
                "streak": streak,
                "totalActiveDays": total_active_days,
                "calendarCounts": calendar_counts,
                "recentAcceptedSubmissions": processed_subs
            }

            sub_count = len(processed_subs)
            total_submissions_collected += sub_count
            success_count += 1
            print(f"OK ({len(recent_subs_raw)} recent AC subs)")

        except Exception as e:
            fail_count += 1
            print(f"FAILED: {e}")

        # Polite API delay between students
        time.sleep(0.5)

    # Save daily submissions dataset
    with open(daily_file_path, "w", encoding="utf-8") as f:
        json.dump(daily_dataset, f, indent=2, ensure_ascii=False)

    exec_time = time.time() - start_time
    print("\nCollection completed.")
    print("--------------------------------------------------")
    print(f"Students processed         : {total_students}")
    print(f"Successful students        : {success_count}")
    print(f"Failed students            : {fail_count}")
    print(f"Accepted subs collected    : {total_submissions_collected}")
    print(f"Unique problem slugs cached: {len(problem_cache)}")
    print(f"Execution time             : {exec_time:.2f} seconds")
    print(f"Output saved to            : {daily_file_path}")
    print(f"Problem cache saved to     : {CACHE_FILE}")

if __name__ == "__main__":
    main()
